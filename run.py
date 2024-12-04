"""
This module provides functionality for scraping event data
with BeutifulSoup from Eventbrite, processing and storing the
data in MongoDB. Events can be viewed in console or there are
functions for uploading user collected event data to CSV/Excel
or generate visualizations with collected data to Google Cloud Storage
so they can be accessed by the user with links from the console.

Modules and librarys used are below:
"""
import base64
import calendar
import csv
import itertools
import math
import os
import re
import sys
import threading
import time
from collections import Counter
from datetime import datetime
import matplotlib.pyplot as plt
import openpyxl
import requests
from bs4 import BeautifulSoup
from dateutil import parser
from dotenv import load_dotenv
from geopy.distance import geodesic
from google.cloud import storage
from openpyxl.utils import get_column_letter
from pymongo.mongo_client import MongoClient
from pymongo.server_api import ServerApi
from pymongo.errors import ConnectionFailure, OperationFailure

load_dotenv()
encoded_json = os.getenv('GOOGLE_CREDENTIALS')
decoded_json = base64.b64decode(encoded_json).decode()
# Decode the required credentials from the environment variable

with open('/tmp/service_account.json', 'w', encoding='utf-8') as f:
    f.write(decoded_json)
    # Write the decoded credentials to a JSON file

os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = '/tmp/service_account.json'
# Use the decoded credentials to authenticate with Google Cloud

# Hashtable to cache recently searched events
cache = {}

uri = os.getenv('MONGO_URI')
# Create a new client and connect to the server
client = MongoClient(uri, server_api=ServerApi('1'))
# Send a ping to confirm a successful connection
try:
    client.admin.command('ping')
    print("Pinged your deployment. You successfully connected to MongoDB!")
except ConnectionFailure as e:
    print(f'Connection error: {e}')
except OperationFailure as e:
    print(f'Operation failure: {e}')

db = client['Event_Hoarder']
collection = db['Event_Data']
# MongoDB database and collection

# Directory to save Excel and CSV files
UPLOAD_FOLDER = 'data_visuals'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

stored_urls = []
processed_files = set()


def upload_to_gcs(bucket_name, source_file_name, destination_blob_name):
    """
    Uploads a file to a googlle cloud storage bucket.

    Args:
        bucket_name (str): Name of the GCS bucket.
        source_file_name (str): The path to file for upload.
        destination_blob_name (str): Name of the file sent to GCS.

    Returns:
        str: Public GCS URL of the uploaded file.
    """
    storage_client = storage.Client()
    # Create a new client
    bucket = storage_client.bucket(bucket_name)
    # Get the bucket by name
    blob = bucket.blob(destination_blob_name)
    # A blob is a file in Google Cloud Storage
    blob.upload_from_filename(source_file_name)
    # Upload the file to the bucket
    url = blob.public_url
    # Get the public URL of the uploaded file
    print(f"File uploaded to {url}")
    stored_urls.append(url)
    # Append the URL to the stored_urls list for viewing
    return url


def delete_all_files_in_gcs(bucket_name):
    """
    Removes all files from a Google Cloud Storage bucket.

    Args:
        bucket_name (str): The name of the GCS bucket to delete files from.
    """
    storage_client = storage.Client()
    bucket = storage_client.bucket(bucket_name)
    blobs = bucket.list_blobs()

    # List all blobs in the bucket
    for blob in blobs:
        blob.delete()
        print(f"Deleted {blob.name}")

    # Confirm all blobs are deleted
    remaining_blobs = list(bucket.list_blobs())
    if not remaining_blobs:
        print('All files deleted from Google Cloud Storage.')
    else:
        print('Some files could not be deleted.')

    processed_files.clear()


class Spinner:
    """
    A class to create a spinner (loader) that runs in a separate thread.
    """
    def __init__(self, message='Loading...'):
        """
        Initializes the Spinner object with a message and a spinner.

        Args:
            message (str, optional): Defaults to 'Loading...'.
        """
        self.message = message
        self.spinner = itertools.cycle(['|', '/', '-', '\\'])
        self.stop_running = threading.Event()
# This sets up the message and spinner, and creates a stop_running event that
# will be used to stop the spinner.

    def start(self):
        """
        This starts a new thread that will run the _spin method.
        """
        threading.Thread(target=self._spin).start()

    def _spin(self):
        """
        This method will run as long as stop_running event is not set.
        """
        while not self.stop_running.is_set():
            # This method will run as long as stop_running event is not set.
            sys.stdout.write(f'\r{self.message} {next(self.spinner)}')
            # \r is a carriage return, which moves the cursor to the beginning
            # of the line. This allows the spinner to overwrite itself on the
            # same line, next(self.spinner) gets next character in the spinner.
            sys.stdout.flush()
            # Immediately flushes to the console to show the spinner.
            time.sleep(0.1)
            sys.stdout.write('\b')
            # \b is a backspace, which moves the cursor back one character
            # so a new character can be written.

    def stop(self):
        """
        This method sets the stop_running event, which will stop the spinner.
        """
        self.stop_running.set()
        # Thread is stopped
        sys.stdout.write('\r' + ' ' * (len(self.message) + 2) + '\r')
        # Clean up the spinner by overwriting it with spaces and
        # moving the cursor back to the beginning of the line.
        sys.stdout.flush()


def view_data_files():
    """
    View files uploaded to Google Cloud Storage from global list stored_urls.
    """
    if not stored_urls:
        print('\n-------------------------------------\nNo files uploaded yet.\
            \n-------------------------------------')
    else:
        for url in stored_urls:
            print(url)
    input('DO NOT USE CTRL-C TO COPY\nGo back to the main menu? Press Enter to'
          'continue.')
    main()
    return


def check_and_delete_old_events():
    """
    Check for events in the global collection with start dates that have
    already passed and delete them.
    """
    current_date = datetime.now().date()
    events = collection.find({})

    for event in events:
        unique_id = event.get('url', 'N/A')
        if unique_id == 'N/A':
            continue

        start_date = event.get('event_date_time', 'N/A')

        try:
            checked_start_date = datetime.strptime(
                start_date, '%Y-%m-%d %H:%M:%S').date()
        except ValueError:
            collection.delete_one({'url': unique_id})
            continue

        if current_date > checked_start_date:
            collection.delete_one({'url': unique_id})


check_and_delete_old_events()


def save_to_mongodb(search_key, collected_events):
    """
    Save user viewed events to MongoDB.

    Args:
        search_key (str): The phrase the user used to search for events.
        collected_events (list of dict): The user's collected events.
    """
    for event in collected_events:
        if isinstance(event, dict):
            unique_id = event.get('url', 'N/A')
            if unique_id == 'N/A':
                continue

            event_data = {
                'search_key': search_key,
                'url': unique_id,
                'name': event.get('name', 'N/A'),
                'location': event.get('location', 'N/A'),
                'event_date_time': event.get('event_date_time', 'N/A'),
                'show_date_time': event.get('show_date_time', 'N/A'),
                'summary': event.get('summary', 'N/A'),
                'event_price': event.get('event_price', 'N/A'),
                'event_organiser_name': event.get(
                    'event_organiser_name', 'N/A'),
                'event_organiser_link': event.get(
                    'event_organiser_link', 'N/A'),
            }

            collection.update_one(
                {'url': unique_id}, {'$set': event_data}, upsert=True)
        else:
            print(f"Skipping invalid event: {event}")


def save_to_csv(events):
    """
    Save events in the mongodb collection to a CSV file.

    Args:
        events (list of dict): The users collected events.
    """
    directory = 'data_visuals'
    file_name = os.path.join(directory, 'collected_events.csv')
    file_exists = os.path.exists(file_name)

    fields = [
        'name', 'location', 'show_date_time', 'event_price', 'summary',
        'url', 'event_organiser_name', 'event_organiser_link'
        ]

    with open(file_name, 'a', newline='', encoding='utf-8') as file:
        writer = csv.DictWriter(file, fieldnames=fields)

        if not file_exists:
            writer.writeheader()

        for event in events:
            filtered_event = {
                field: event[field] for field in fields if field in event}
            # this will only include the fields that are in the fields
            # list above, so that fields that only have a programmatic purpose
            # are not included in the CSV it works by creating a new
            # dictionary with only the fields that are in the fields list
            writer.writerow(filtered_event)

    print(f'\n-------------------------------------\nEvents saved to'
          f'{file_name}, download/view from the main menu.'
          f'\n-------------------------------------')
    upload_to_gcs(
        'data-visuals-serving', file_name, os.path.basename(file_name))
    # os.path.basename() example: 'data_visuals/collected_events.csv'
    # -> 'collected_events.csv'
    return


def save_to_excel(events, filename='data_visuals/events_data.xlsx'):
    """
    Save events in the mongodb collection to an Excel file.

    Args:
        events (list of dict): The users collected events.
        filename (str, optional): The filename for the Excel file,
                                Defaults to 'data_visuals/events_data.xlsx'.
    """
    filename = check_file_unique(filename)
    workbook = openpyxl.Workbook()
    # Create a new Excel workbook
    sheet = workbook.active
    # active means the first sheet in the workbook
    sheet.title = 'Events Data'
    # Set the title of the sheet to 'Events Data'
    headers = [
        'Event Name', 'Date', 'Location', 'Price',
        'Summary', 'URL', 'Organiser', 'Organiser Link']
    column_widths = [71, 58, 111, 12, 81, 140, 71, 140]
    # Set the column headers and their widths

    for col_num, (header, width) in enumerate(zip(headers, column_widths), 1):
        # for every col_num and its corresponding header and width, enumerate
        # will return the index and value of each column, and then associate
        # the declared headers and column_widths stated above with the index
        col_letter = get_column_letter(col_num)
        # get_column_letter() is a built in function from openpyxl that
        # returns the letter of the specified column,
        # example: 1 -> 'A', 2 -> 'B'
        sheet[f'{col_letter}1'] = header
        # Will set the headers stated above in the first row of the sheet
        # iterating through columns A, B, C, D, E, F
        sheet.column_dimensions[col_letter].width = width
        # column_dimensions is a dictionary that stores the width of each
        # column, the width is set to the width in the column_widths list

    for row_num, event in enumerate(events, 2):
        sheet[f'A{row_num}'] = event.get('name', 'N/A')
        sheet[f'B{row_num}'] = event.get('show_date_time', 'N/A')
        sheet[f'C{row_num}'] = event.get('location', 'N/A')
        sheet[f'D{row_num}'] = event.get('event_price', 'N/A')
        sheet[f'E{row_num}'] = event.get('summary', 'N/A')
        sheet[f'F{row_num}'] = event.get('url', 'N/A')
        sheet[f'G{row_num}'] = event.get('event_organiser_name', 'N/A')
        sheet[f'H{row_num}'] = event.get('event_organiser_link', 'N/A')

    workbook.save(filename)
    print(f'\n-------------------------------------\nEvents saved to'
          f'{filename}, download/view from the main menu.'
          f'\n-------------------------------------')
    upload_to_gcs('data-visuals-serving', filename, os.path.basename(filename))
    return


def parsed_scraped_date(date_time):
    """
    Turn the user friendley (readable) event date and time into a
    computer readable datetime object.

    Args:
        date_time (str): User friendly date and time string.

    Returns:
        string: A date_time string in the format '%Y-%m-%d %H:%M:%S'.
    """
    if 'No date and time available' in date_time or not date_time.strip():
        return 'N/A'
    # If the date_time is not available, return 'N/A'
    replacements = {
        'Monday': '',
        'Mon': '',
        'Tuesday': '',
        'Tue': '',
        'Wednesday': '',
        'Wed': '',
        'Thursday': '',
        'Thu': '',
        'Friday': '',
        'Fri': '',
        'Saturday': '',
        'Sat': '',
        'Sunday': '',
        'Sun': '',
        'Starts on': '',
        'GMT': '',
        'GMT+1': '',
        'January': 'Jan',
        'February': 'Feb',
        'March': 'Mar',
        'April': 'Apr',
        'May': 'May',
        'June': 'Jun',
        'July': 'Jul',
        'August': 'Aug',
        'September': 'Sep',
        'October': 'Oct',
        'November': 'Nov',
        'December': 'Dec',
        'pm': 'PM',
        'am': 'AM',
        '·': '',
        ' - ': ' ',
        '+1': '',
    }

    for key, value in replacements.items():
        date_time = date_time.replace(key, value)
    # Replace the long names of the days and months with their abbreviations
    # Also replace other unwanted strings

    # Remove any extra spaces and commas
    date_time = ' '.join(date_time.split()).replace(',', '')

    # Identify and correct multiple months in the date_time
    month_abbrs = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    month_count = {month: date_time.count(month) for month in month_abbrs}
    # Counts the occurrences of each month abbreviation in the date_time
    months_found = [month for month, count in month_count.items() if count > 0]

    if len(months_found) > 1:
        # Replace first month witht he last one specified
        first_month = months_found[0]
        last_month = months_found[-1]
        date_time = date_time.replace(first_month, last_month, 1)

    # For cases where users enter date ranges, only take the first date
    date_time_parts = date_time.split(' ')
    # Only take the first 5 parts of the date_time
    if len(date_time_parts) > 3:
        date_time = ' '.join(date_time_parts[:3])
    # If the length of the date_time is less than 3, return the date_time as is
    # If the length of the date_time is greater than 3,
    # only take the first 5 parts of the date_time
    # parts example: ['2024-10-19', '16:30:00']

    try:
        # Use dateutil.parser to parse the date
        dt = parser.parse(date_time, fuzzy=True)
        # example: 2024-10-19 16:30:00
    except ValueError:
        # parser.parse will try to parse the date and time from the string
        # fuzzy=True allows for more flexibility in the date format
        parts = date_time.split()
        # example ['2024-10-19', '16:30:00']
        if len(parts) >= 2:
            try:
                year, month, day = map(int, parts[0].split('-'))
                last_valid_day = calendar.monthrange(year, month)[1]

                if day > last_valid_day:
                    day = last_valid_day

                corrected_date = f'{year}-{month}-{day} {parts[1]}'
                dt = parser.parse(corrected_date, fuzzy=True)
            except ValueError:
                return 'N/A'

    # Format the datetime object into the desired format
    formatted_date = dt.strftime('%Y-%m-%d %H:%M:%S')

    return formatted_date


def display_events(events, start_index, end_index, user_selection, search_key):
    """
    Display events to the user in a readable, friendley format.

    Args:
        events (list of dict): Scraped event data.
        start_index (int): The start index of the events to display.
        end_index (int): The end index of the events to display.
        user_selection (str): A backend string to determine what to do
        search_key (str): The phrase the user used to search for events.

    Returns:
        List of dict: The events displayed to the user.
    """
    collected_events = events[start_index:end_index]
    for data in collected_events:
        if isinstance(data, dict):
            show_date_time = data.get('show_date_time',
                                      'No date and time available')
            summary = data.get('summary', '')
            truncated_summary = (
                summary[:120] + '...' if len(summary) > 120 else summary)
            if show_date_time != 'No date and time available':
                print(f'-------------------------------------\n{data["name"]},'
                      f'\n{data["location"]}\n{data["show_date_time"]}'
                      f'\nPrice: {data["event_price"]}'
                      f'\nSummary: {truncated_summary}'
                      f'\nEvent URL: {data["url"]}'
                      f'\nOrganiser: {data["event_organiser_name"]}'
                      f'\nOrganiser\'s Link: {data["event_organiser_link"]}'
                      f'\n-------------------------------------')
            else:
                continue  # Skip invalid event data
        else:
            continue  # Skip invalid event data

    if user_selection == 'data-manipulation':
        return print('-------------------------------------'
                     '\nEvents displayed in relevance bottom to top.')
    elif user_selection in ('eventbrite', 'eventbrite_top',
                            'eventbrite_top_no_category'):
        save_to_mongodb(search_key, collected_events)

    # Cache the events in the hashtable
    cache[search_key] = events


def scrape_eventbrite_events(location, day, product, page_number,
                             start_date, end_date):
    """
    Scrape events from Eventbrite using a quick search.

    Args:
        location (str): The users given location. (optional)

        day (str): A string representing eg. 'today',
                'tomorrow', 'this-weekend'. (optional)

        product (str): The type of event the user is looking for. (optional)

        page_number (int): The page number of the search
                        results for backend

        start_date (str): The users given date to start
                        the search from. (optional)

        end_date (str): The users given date to end the search. (optional)

    Returns:
        tuple: A tuple containing:
            - event_data (list): A list of dictionaries,
                each representing an event.
            - tags_counter (Counter): A Counter object counting
                the occurrences of tags.
            - more_events_check (bool): A boolean indicating if there
                are more events to fetch.
    """
    url = (
        f'https://www.eventbrite.com/d/united-kingdom--{location}/events--'
        f'{day}/{product}/?page={page_number}&start_date={start_date}&end_date'
        f'={end_date}'
    )
    page = requests.get(url, timeout=20)
    soup = BeautifulSoup(page.content, 'html.parser')
    events = soup.find_all('a', class_='event-card-link')

    event_data = []
    tags_counter = Counter()
    seen_urls = set()

    for event in events:
        event_url = event['href']
        if event_url in seen_urls:
            continue
        seen_urls.add(event_url)

        event_info = {
            'name': event.get('aria-label', '').replace('View', '').strip(),
            'url': event_url
        }

        page_detail = requests.get(event_url, timeout=20)
        page_detail_soup = BeautifulSoup(page_detail.content, 'html.parser')

        price_div = page_detail_soup.find('div',
                                          class_="conversion-bar__panel-info")
        event_price = price_div.get_text(strip=True) if price_div else 'Free'

        location_div = page_detail_soup.find('div',
                                             class_='location-info__address')
        if location_div:
            event_location = location_div.get_text(separator=' ', strip=True)
            # Code with help from Co-Pilot
            event_location = re.sub(r'Show map$', '', event_location)
            # End of Co-Pilot code, this code removes the 'Show map' text
            # from the location even if its preceeded by another word,
            # the issue was "United KingdomShow map"
        else:
            location_div = page_detail_soup.find(
                'div', class_='location-info__address-text')
            if location_div:
                event_location = location_div.get_text(
                    separator=' ', strip=True)

        summary_div = page_detail_soup.find('div', class_='eds-text--left')
        event_summary = ''
        if summary_div:
            p_elements = summary_div.find_all('p')
            summary_texts = [p_element.get_text(
                separator=' ', strip=True) for p_element in p_elements]
            event_summary = ' '.join(summary_texts)
        elif not summary_div:
            summary = page_detail_soup.find('p', class_='summary')
            event_summary = summary.get_text(
                strip=True) if summary else 'No summary available'

        date_time = page_detail_soup.find(
            'span', class_='date-info__full-datetime')
        event_date_time = date_time.get_text(
            strip=True) if date_time else 'No date and time available'

        event_organiser_name = 'No organiser available'
        event_organiser_link = None

        event_organiser_divs = page_detail_soup.find_all(
            'div', class_='descriptive-organizer-info-heading-signal-container'
            )

        for div in event_organiser_divs:
            organiser_link = div.find(
                'a', class_='descriptive-organizer-info-mobile__name-link')
            if organiser_link:
                event_organiser_name = organiser_link.get_text(strip=True)
                event_organiser_link = organiser_link['href']

        date_parsed = parsed_scraped_date(event_date_time)

        tags = page_detail_soup.find_all(
            'a', class_='tags-link listing-tag eds-l-mar-top-4'
            ' eds-text-bs eds-text--center')
        for tag in tags:
            tags_counter[tag.get_text(strip=True)] += 1

        event_info.update({
            'location': event_location,
            'show_date_time': event_date_time,  # User clear date
            'event_date_time': date_parsed,
            'summary': event_summary,
            'event_price': event_price,
            'event_organiser_name': event_organiser_name,
            'event_organiser_link': event_organiser_link,
        })

        event_data.append(event_info)

    more_events_check = len(events) > 0  # Check if more events on next page

    return event_data, tags_counter, more_events_check


def scrape_eventbrite_categories(location, category_slug, day,
                                 page_number, start_date, end_date):
    """
    Scrape events from Eventbrite using a set category search.

    Args:
        location (string): The users given location. (optional)

        category_slug (string): A slug representing the chosen category.

        day (string): A string representing eg. 'today',
        'tomorrow', 'this-weekend'. (optional)

        page_number (interger): The page number of the search
        results for backend

        start_date (string): The users given date to start
        the search from. (optional)

        end_date (string): The users given date to end the search. (optional)

    Returns:
        tuple: A tuple containing:
        - event_data (list): A list of dictionaries,
            each representing an event.
        - tags_counter (Counter): A Counter object counting
            the occurrences of tags.
        - more_events_check (bool): A boolean indicating if
            there are more events to fetch.
    """
    url = (
        f'https://www.eventbrite.com/d/united-kingdom--{location}/'
        f'{category_slug}--events--{day}/?page={page_number}&start_date='
        f'{start_date}&end_date={end_date}'
    )
    page = requests.get(url, timeout=20)
    soup = BeautifulSoup(page.content, 'html.parser')
    events = soup.find_all('a', class_='event-card-link')

    event_data = []
    tags_counter = Counter()
    seen_urls = set()

    for event in events:
        event_url = event['href']
        if event_url in seen_urls:
            continue
        seen_urls.add(event_url)

        event_info = {
            'name': event.get('aria-label', '').replace('View', '').strip(),
            'url': event_url
        }

        page_detail = requests.get(event_url, timeout=20)
        page_detail_soup = BeautifulSoup(page_detail.content, 'html.parser')

        price_div = page_detail_soup.find(
            'span', class_="eds-text-bm eds-text-weight--heavy")
        event_price = price_div.get_text(strip=True) if price_div else 'Free'

        location_div = page_detail_soup.find(
            'div', class_='location-info__address')
        if location_div:
            event_location = location_div.get_text(separator=' ', strip=True)
            # Code with help from Co-Pilot
            event_location = re.sub(r'Show map$', '', event_location)
            # End of Co-Pilot code, this code removes the 'Show map'
            # text from the location even if its preceeded by another word,
            # the issue was "United KingdomShow map"
        else:
            location_div = page_detail_soup.find(
                'div', class_='location-info__address-text')
            if location_div:
                event_location = location_div.get_text(
                    separator=' ', strip=True)

        summary_div = page_detail_soup.find('div', class_='eds-text--left')
        event_summary = ''
        if summary_div:
            p_elements = summary_div.find_all('p')
            summary_texts = [p_element.get_text(
                separator=' ', strip=True) for p_element in p_elements]
            event_summary = ' '.join(summary_texts)
        elif not summary_div:
            summary = page_detail_soup.find('p', class_='summary')
            event_summary = summary.get_text(
                strip=True) if summary else 'No summary available'

        date_time = page_detail_soup.find(
            'span', class_='date-info__full-datetime')
        event_date_time = date_time.get_text(
            strip=True) if date_time else 'No date and time available'

        event_organiser_name = 'No organiser available'
        event_organiser_link = None

        event_organiser_divs = page_detail_soup.find_all(
            'div', class_='descriptive-organizer-info-heading-signal-container'
            )

        for div in event_organiser_divs:
            organiser_link = div.find(
                'a', class_='descriptive-organizer-info-mobile__name-link')
            if organiser_link:
                event_organiser_name = organiser_link.get_text(strip=True)
                event_organiser_link = organiser_link['href']

        date_parsed = parsed_scraped_date(event_date_time)

        tags = page_detail_soup.find_all(
            'a', class_='tags-link listing-tag eds-l-mar-top-4 eds-text-bs'
            ' eds-text--center')
        for tag in tags:
            tags_counter[tag.get_text(strip=True)] += 1

        event_info.update({
            'location': event_location,
            'show_date_time': event_date_time,
            'event_date_time': date_parsed,
            'summary': event_summary,
            'event_price': event_price,
            'event_organiser_name': event_organiser_name,
            'event_organiser_link': event_organiser_link,
        })

        event_data.append(event_info)

    more_events_check = len(events) > 0

    return event_data, tags_counter, more_events_check


def scrape_eventbrite_top_events(location):
    """
    Scrape events from Eventbrite using a given
    location to find top events there.

    Args:
        location (string): The users given location. (optional)

    Returns:
        tuple: A tuple containing:
        - event_data (list): A list of dictionaries,
            each representing an event.
        - tags_counter (Counter): A Counter object counting
            the occurrences of tags.
        - more_events_check (bool): A boolean indicating if
            there are more events to fetch.
    """
    url = f'https://www.eventbrite.co.uk/d/united-kingdom--{location}/events/'
    page = requests.get(url, timeout=20)
    soup = BeautifulSoup(page.content, 'html.parser')
    events = soup.find_all('a', class_='event-card-link')

    event_data = []
    tags_counter = Counter()
    seen_urls = set()

    for event in events:
        event_url = event['href']
        if event_url in seen_urls:
            continue
        seen_urls.add(event_url)

        event_info = {
            'name': event.get('aria-label', '').replace('View', '').strip(),
            'url': event_url
        }

        page_detail = requests.get(event_url, timeout=20)
        page_detail_soup = BeautifulSoup(page_detail.content, 'html.parser')

        price_div = page_detail_soup.find(
            'div', class_="conversion-bar__panel-info")
        event_price = price_div.get_text(strip=True) if price_div else 'Free'

        location_div = page_detail_soup.find(
            'div', class_='location-info__address')
        if location_div:
            event_location = location_div.get_text(separator=' ', strip=True)
            # Code with help from Co-Pilot
            event_location = re.sub(r'Show map$', '', event_location)
            # End of Co-Pilot code, this code removes the 'Show map'
            # text from the location even if its preceeded by another word,
            # the issue was "United KingdomShow map"
        else:
            location_div = page_detail_soup.find(
                'div', class_='location-info__address-text')
            if location_div:
                event_location = location_div.get_text(
                    separator=' ', strip=True)

        summary_div = page_detail_soup.find('div', class_='eds-text--left')
        event_summary = ''
        if summary_div:
            p_elements = summary_div.find_all('p')
            summary_texts = [p_element.get_text(
                separator=' ', strip=True) for p_element in p_elements]
            event_summary = ' '.join(summary_texts)
        elif not summary_div:
            summary = page_detail_soup.find('p', class_='summary')
            event_summary = summary.get_text(
                strip=True) if summary else 'No summary available'

        date_time = page_detail_soup.find(
            'span', class_='date-info__full-datetime')
        event_date_time = date_time.get_text(
            strip=True) if date_time else 'No date and time available'

        event_organiser_name = 'No organiser available'
        event_organiser_link = None

        event_organiser_divs = page_detail_soup.find_all(
            'div', class_='descriptive-organizer-info-heading-signal-container'
            )

        for div in event_organiser_divs:
            organiser_link = div.find(
                'a', class_='descriptive-organizer-info-mobile__name-link')
            if organiser_link:
                event_organiser_name = organiser_link.get_text(strip=True)
                event_organiser_link = organiser_link['href']

        date_parsed = parsed_scraped_date(event_date_time)

        tags = page_detail_soup.find_all(
            'a', class_='tags-link listing-tag eds-l-mar-top-4'
            ' eds-text-bs eds-text--center')
        for tag in tags:
            tags_counter[tag.get_text(strip=True)] += 1

        event_info.update({
            'location': event_location,
            'show_date_time': event_date_time,  # User clear date
            'event_date_time': date_parsed,
            'summary': event_summary,
            'event_price': event_price,
            'event_organiser_name': event_organiser_name,
            'event_organiser_link': event_organiser_link,
        })

        event_data.append(event_info)

    more_events_check = len(events) > 0  # Check if more events on next page

    return event_data, tags_counter, more_events_check


def collection_menu():
    """
    This menu provides the user with options to view
    the events in the collection
    """
    while True:
        print('\n-------------------------------------'
              '\nOn this menu you may view your all your collected events,'
              '\n view recentley searched events, or clear the database.'
              '\nAfter here when you choose how to view the events,'
              '\n you may export to CSV (C) and Excel (E) or perform'
              '\n data tasks with these events (T).'
              '\n-------------------------------------')
        print('\nChoose an option to manipulate events or print to CSV:')
        print('1. View all searched events')
        print('2. View recent searches')
        print('3. Main Menu')
        print('#. Clear Database')
        choice = input('Enter your choice: ').strip()

        if choice == '1':
            view_all_events()
        elif choice == '2':
            search_events_in_collection()
        elif choice == '3':
            print("Going back to the main menu.")
            main()
            return
        elif choice == '#':
            collection.delete_many({})
            print('-------------------------------------\nDatabase cleared'
                  '\n-------------------------------------.')
            main()
            return
        else:
            print('\n-------------------------------------'
                  '\nInvalid choice. Please try again.'
                  '\n-------------------------------------')

# extract_price was built with help from Co-Pilot when asked how could
# I extract the price from the event_price field
# Its suggestion was the re module. As I needed to remove other strings and
# currency symbols from the event_price field


def extract_price(price_str):
    """
    The extract_price function uses a regular expression to find the first
    occurrence of a number in the event_price string.
    If a number is found, it is converted to a float and returned. If no number
    is found, 0.0 is returned. The sort_events function uses the extract_price
    function to extract the numeric part of the event_price before sorting.

    Args:
        price_str (str): The price string to extract the price from.

    Returns:
        Float: The numeric part of the event price string, or 0.0 if no number
                is found.
    """

    match = re.search(r'\d+(\.\d+)?', price_str)
    # \d+(\.\d+)? means match one or more digits followed by an
    # optional decimal point and one or more digits
    if match:
        return float(match.group())
    return 0.0


# End of Co-Pilot code


def check_file_unique(image_path):
    """
    This function splits the image_path into a
    directory and filename, then checks if the filename already exists.
    If so, add +1 to the filename and return the new image_path.

    Args:
        image_path (string): The path to the visualisation image.

    Returns:
        string: If the image already exists, return a new image_path.
    """
    directory, filename = os.path.split(image_path)
    # Example: 'data_visuals/event_count_per_day.png' ->
    # ('data_visuals', 'event_count_per_day.png')
    base, ext = os.path.splitext(filename)
    # Example: 'event_count_per_day.png' -> ('event_count_per_day', '.png')
    counter = 1

    while os.path.exists(image_path):
        image_path = os.path.join(directory, f'{base}_{counter}{ext}')
        # While theres already a file with the same name,
        # add a counter to the filename
        counter += 1

    return image_path


def compare_events(events):
    """
    A menu to compare collected events with different
    analytics and visualisations.

    Args:
        events (list of dict): The user's collected events.
    """
    if len(events) < 2:
        print('Not enough events to compare.')
        return

    while True:
        print('\n-------------------------------------'
              '\nAfter you select an option, your selection will be displayed'
              ' \nor saved as a visualisation, that can be viewed/downloaded'
              ' \nfrom a provided link in the main menu with option 5'
              '\n-------------------------------------')
        print('\nWhat would you like to compare?')
        print('1. Average price of events')
        print('2. Median price of events')
        print('3. Event count per day')
        print('4. Event count per month')
        print('5. Event price distribution')
        print('6. Event dates over time')
        print('7. Main Menu')
        choice = input('Enter your choice: ').strip()

        if choice == '1':
            spinner = Spinner('Processing...')
            spinner.start()
            try:
                price = [extract_price(event.get(
                    'event_price', '0')) for event in events if event.get(
                        'event_price', 'N/A').lower() not in [
                            'sold out', 'free', 'donation']]
                # Extract the price from the event_price field for each event,
                # not including 'sold out', 'free', and 'donation' events
                result = sum(price) / len(price) if price else 0
                # Calculate the average price of the events, sum the prices
                # and divide by the number of prices
                floored_result = math.floor(result)
                # Round down the result to the nearest whole number
                print(f'\nThe average price of events is: £{floored_result}')
            finally:
                spinner.stop()
        elif choice == '2':
            spinner = Spinner('Processing...')
            spinner.start()
            try:
                price = [extract_price(event.get(
                    'event_price', '0')) for event in events if event.get(
                        'event_price', 'N/A').lower() not in [
                            'sold out', 'free', 'donation']]
                # Extract the price from the event_price field for each event,
                # not including 'sold out', 'free', and 'donation' events
                if price:
                    price.sort()
                    # Sort the list of prices in ascending order
                    n = len(price)
                    # Get length of list
                    mid = n // 2
                    # Find the middle index
                    if n % 2 == 0:
                        result = (price[mid - 1] + price[mid]) / 2
                        # If the list length is even, the median is
                        # the average of the two middle elements
                    else:
                        result = price[mid]
                        # If the list length is odd,
                        # the median is the middle element
                else:
                    result = 0
                    # If empty list, return 0
                print(f'\nThe median price of events is: £{result}')
            finally:
                spinner.stop()
        elif choice == '3':
            spinner = Spinner('Processing...')
            spinner.start()
            try:
                event_days = [datetime.strptime(
                    event['event_date_time'], '%Y-%m-%d %H:%M:%S').strftime(
                        '%Y-%d') for event in events if event.get(
                            'event_date_time', 'N/A') != 'N/A']
                # List comprehension to extract the day and year from the
                # event_date_time field for each event, the date is expected
                # to be in the format 'YYYY-MM-DD HH:MM:SS' and
                # is parsed to a datetime object. The second argument formats
                # the datetime object to 'YYYY-DD' to be used to group
                # the events by day
                days_counts = Counter(event_days)
                # Use the Counter class to count the occurrences of each day,
                # works by setting a dictionary key with each collected day
                # and incrementing the value each time the day is found
                days, counts = zip(*sorted(days_counts.items()))
                # Take the dictionary days_counts and retrieve its items as a
                # list of tuples with items(), then sort the tuples by the keys
                # (days) sorted(), take the sorted list of tuples and unpacks
                # them into two lists, one for keys and one for values zip(*).
                # The splat operator * allowed zip() to take the list of tuples
                # and unpack them into two separate lists, without it zip()
                # would return one list of tuples
                plt.bar(days, counts)
                # Create a bar chart with the days on the x-axis
                # and the counts on the y-axis
                plt.title('Events By Day')
                plt.xlabel('Day')
                plt.ylabel('Number of Events')
                plt.xticks(rotation=45)
                # Rotate the x-axis labels by 45 degrees for better readability
                plt.tight_layout()
                # Automatic padding
                image_path = 'data_visuals/event_count_per_day.png'
                image_path = check_file_unique(image_path)
                plt.savefig(image_path)
                # Save the plot as an image
                plt.close()
                upload_to_gcs(
                    'data-visuals-serving', image_path, os.path.basename(
                        image_path))
                print(f'\n-------------------------------------'
                      f'\nEvent count per day saved as {image_path},'
                      f' download/view from the main menu.'
                      f'\n-------------------------------------')
            finally:
                spinner.stop()
        elif choice == '4':
            spinner = Spinner('Processing...')
            spinner.start()
            try:
                event_months = [datetime.strptime(
                    event['event_date_time'], '%Y-%m-%d %H:%M:%S').strftime(
                        '%Y-%m') for event in events if event.get(
                            'event_date_time', 'N/A') != 'N/A']
                # List comprehension to extract the month and year from the
                # event_date_time field for each event, the date is expected
                # to be in the format 'YYYY-MM-DD HH:MM:SS' and is parsed to a
                # datetime object. The second argument formats the datetime
                # object to 'YYYY-MM' to be used to group the events by month
                months_counts = Counter(event_months)
                # Use the Counter class to count the occurrences of each month,
                # works by setting a dictionary key with each collected month
                # and incrementing the value each time the month is found
                months, counts = zip(*sorted(months_counts.items()))
                # Take the dictionary months_counts and retrieve its items as
                # a list of tuples with items(), then sort the tuples by the
                # keys (months) sorted(), take the sorted list of tuples
                # and unpacks them into two lists, one for keys and one for
                # values zip(*). The splat operator * allowed zip() to take
                # the list of tuples and unpack them into two separate lists,
                # without it zip() would return one list of tuples
                plt.bar(months, counts)
                # Create a bar chart with the months on the x-axis and
                # the counts on the y-axis
                plt.title('Events By Month')
                plt.xlabel('Month')
                plt.ylabel('Number of Events')
                plt.xticks(rotation=45)
                # Rotate the x-axis labels by 45 degrees for better readability
                plt.tight_layout()
                # Automatic padding
                image_path = 'data_visuals/event_count_per_month.png'
                image_path = check_file_unique(image_path)
                plt.savefig(image_path)
                # Save the plot as an image
                plt.close()
                upload_to_gcs(
                    'data-visuals-serving', image_path, os.path.basename(
                        image_path))
                print(f'\n-------------------------------------'
                      f'\nEvent count per month saved as {image_path},'
                      f' download/view from the main menu.'
                      f'\n-------------------------------------')
            finally:
                spinner.stop()
        elif choice == '5':
            spinner = Spinner('Processing...')
            spinner.start()
            try:
                price = [extract_price(event.get(
                    'event_price', '0')) for event in events if event.get(
                        'event_price', 'N/A').lower() not in [
                            'sold out', 'free', 'donation']]
                # List comprehension to extract the prices from the events,
                # not including 'sold out', 'free', and 'donation' events,
                # then calling the extract_price function to extract
                # the numeric part of the price
                plt.hist(price, bins=20, edgecolor='black')
                # A bin is a range of values that is used to group the data,
                # the bins argument specifies the number of bins to use,
                # the edgecolor argument specifies the
                # color of the edges of the bars
                plt.title('Event Price Distribution')
                plt.xlabel('Price (£)')
                plt.ylabel('Frequency')
                image_path = 'data_visuals/event_price_distribution.png'
                image_path = check_file_unique(image_path)
                plt.savefig(image_path)
                plt.close()
                upload_to_gcs(
                    'data-visuals-serving', image_path, os.path.basename(
                        image_path))
                print(f'\n-------------------------------------'
                      f'\nEvent price distribution saved as {image_path},'
                      f' download/view from the main menu.'
                      f'\n-------------------------------------')
            finally:
                spinner.stop()
        elif choice == '6':
            spinner = Spinner('Processing...')
            spinner.start()
            try:
                event_dates = [
                    datetime.strptime(event['event_date_time'],
                                      '%Y-%m-%d %H:%M:%S').date()
                    for event in events
                    if event.get('event_date_time', 'N/A') != 'N/A']
                # Grab all the event dates from the event_date_time field,
                # parse the date and time to a datetime object, extract date
                date_counts = Counter(event_dates)
                # Count each occurrence of the different dates
                dates, counts = zip(*sorted(date_counts.items()))
                # Sort the dates and counts, then unpack them into two lists,
                # it works by taking the date_counts Counter dictionary with
                # items() which returns the key-value pairs as a list of
                # tuples, then sorted() sorts the tuples by the keys (dates),
                # then zip() takes the sorted list of tuples and unpacks them
                # into two separate lists using the splat operator *
                plt.plot(dates, counts)
                # Create a line plot with the dates on the x-axis
                # and the counts on the y-axis
                plt.title('Events Over Time')
                plt.xlabel('Date')
                plt.ylabel('Number of Events')
                plt.xticks(rotation=45)
                # Rotate the labels on the x-axis by 45 degrees, readability
                plt.tight_layout()
                # Add padding to the plot
                image_path = 'data_visuals/event_dates_over_time.png'
                image_path = check_file_unique(image_path)
                plt.savefig(image_path)
                plt.close()
                upload_to_gcs(
                    'data-visuals-serving', image_path, os.path.basename(
                        image_path))
                print(f'\n-------------------------------------'
                      f'\nEvent dates over time saved as {image_path},'
                      f' download/view from the main menu.'
                      f'\n-------------------------------------')
            finally:
                spinner.stop()
        elif choice == '7':
            print('Returning to the main menu.')
            main()
            return
        else:
            print('\n-------------------------------------'
                  '\nInvalid choice. Please try again.'
                  '\n-------------------------------------')
            continue


def get_coordinates(location, api_key):
    """
    Get the latitude and longitude coordinates of a location using the
    given location string by the user and use the Google Maps Geocoding API
    to return the coordinates as a tuple.

    Args:
        location (str): The users given location.
        api_key (str): My API key for the Google Maps Geocoding API.

    Returns:
        tuple: A tuple containing the latitude and longitude coordinates.
    """
    url = (
        f'https://maps.googleapis.com/maps/api/geocode/json?address={location}'
        f'&key={api_key}')
    # Construct the URL for the Google Maps Geocoding API
    response = requests.get(url, timeout=20)
    # Send a GET request to the URL
    if response.status_code == 200:
        data = response.json()
    # If the response is successful, turn the response into a JSON object
        if data['status'] == 'OK':
            # Check if the status in the JSON object is 'OK'
            location = data['results'][0]['geometry']['location']
            return (location['lat'], location['lng'])
    # Extract the latitude and longitude from the JSON object
    # and return them as a tuple
    return None


def find_closest_events(user_location, events, api_key):
    """
    Sort the events by the distance to the user's location.

    Args:
        user_location (str): The users given location.
        events (list of dict): The users collected events.
        api_key (str): My API key for the Google Maps Geocoding API.

    Returns:
        list of dict: A sorted list of event dictionaries from bottom to top
        in relavance to the closest to user's location.
    """
    user_coordinates = get_coordinates(user_location, api_key)
    # Get the coordinates of the user's location
    if not user_coordinates:
        print('\n-------------------------------------'
              '\nUser location could not be geocoded.'
              '\n-------------------------------------')
        return []

    def distance_to_user(event):
        event_coordinates = get_coordinates(event['location'], api_key)
        # Get the coordinates of the event's location
        if event_coordinates:
            return geodesic(user_coordinates, event_coordinates).miles
        # The geopy library's geodesic function calculates the distance
        # between two points on the Earth's surface using the geodesic
        # distance, which is more accurate than the haversine formula, as it
        # accounts for the earths ellipsoidal shape, it returns the distance
        # in miles. Source -
        # https://www.askpython.com/python/examples/find-distance-between-two-geo-locations
        return float('inf')
        # If fails return infinity as this is a easy way to sort the
        # events with no coordinates to the end of the list

    events_with_coordinates = [
        event for event in events if event.get('location')]
    # Filter out events with no location
    events_with_coordinates.sort(key=distance_to_user)
    # Sort the events by using the distance_to_user function as the key

    return events_with_coordinates[::-1]
    # Return the sorted events in reverse order


def sort_events(events):
    """
    A menu to sort the collected events by different criteria.

    Args:
        events (list of dict): The user's collected events.
    """
    if len(events) < 2:
        print('Not enough events to sort.')
        return

    while True:
        print('\n-------------------------------------'
              '\nAfter you select an option, your selection will be displayed'
              ' \nor saved as a visualisation, that can be viewed/downloaded'
              ' \nfrom a provided link in the main menu with option 5'
              '\n-------------------------------------')
        print('\nWhat would you like to sort?')
        print('1. Free events')
        print('2. Cheapest events')
        print('3. Most expensive events')
        print('4. Events happening soon')
        print('5. Closest distance events')
        print('6. Main Menu')
        choice = input('Enter your choice: ').strip()

        spinner = Spinner("Sorting events...")
        spinner.start()

        try:
            # Lambda functions are a powerful tool for writing concise,
            # one-off functions, especially useful in situations like sorting,
            # filtering, and mapping.
            if choice == '1':
                free_events = [event for event in events if event.get(
                    'event_price', '').lower() in ['free', 'donation']]
            # Filter in events with event_price of 'free' and 'donation' with
            # list comprehension, if not found, return an empty list
                display_events(free_events[::-1], 0, len(free_events),
                               'data-manipulation', 'None')
            # Display the sorted events from bottom to top
            elif choice == '2':
                cheap_events = [event for event in events if event.get(
                    'event_price', '').lower() not in ['sold out', 'free']]
            # Filter out events with event_price of 'free' and 'sold out' with
            # list comprehension, if not found, return an empty list
                cheap_events_sorted = sorted(
                    cheap_events, key=lambda event: extract_price(event.get(
                        'event_price', '0')))
            # Sort the remaining events in order based on event_price, the key
            # argument specifies a custom sorting function for the sorted()
            # method, and extracts a comparison key from each element.
            # The lambda function is given as the key argument to the sorted()
            # method, it takes argument x that represents each element in the
            # list, and extracts the price from the event_price field
            # with help from the extract_price function. The sorted() method
            # will sort the events in ascending order based on extracted price.
                display_events(cheap_events_sorted[::-1], 0, len(
                    cheap_events_sorted), 'data-manipulation', 'None')
            # Display the sorted events from bottom to top,
            # ::-1 is used to reverse the list
            elif choice == '3':
                paid_events = [event for event in events if event.get(
                    'event_price', '').lower() not in ['free', 'donation']]
            # Filter out events with event_price of 'free' and 'donation' with
            # list comprehension, if not found, return an empty list
                expensive_events_sorted = sorted(
                    paid_events, key=lambda event: extract_price(event.get(
                        'event_price', '0')), reverse=True)
            # Sort the remaining events in reverse order based on event_price,
            # the key argument specifies a custom sorting function for the
            # sorted() method, and extracts a comparison key from each element.
            # The lambda function is given as the key argument to the sorted()
            # method, it takes argument x that represents each element in the
            # list, and extracts the price from the event_price field
            # with help from the extract_price function. The sorted() method
            # will sort the events in decending order based on extracted price.
                display_events(expensive_events_sorted[::-1], 0, len(
                    expensive_events_sorted), 'data-manipulation', 'None')
            # Display the sorted events from bottom to top ,
            # ::-1 is used to reverse the list
            elif choice == '4':
                current_time = datetime.now()
                soonest_events = []
                for event in events:
                    if event.get('event_date_time'):
                        event_date_time = datetime.strptime(event[
                            'event_date_time'], '%Y-%m-%d %H:%M:%S')
                        # Convert the getted event_date_time
                        # string to a datetime object
                        if event_date_time > current_time:
                            soonest_events.append(event)
                        else:
                            continue
                        soonest_events.sort(
                            key=lambda event: event['event_date_time'])
                        # lambda function is used as argument for sort(),
                        # calling a function on every element in the list,
                        # the lambda function takes an event as an argument
                        # and returns the event_date_time
                        # the sort() then kicks in and sorts the events based
                        # on the event_date_time
                    else:
                        continue
                display_events(soonest_events[::-1], 0, len(soonest_events),
                               'data-manipulation', 'None')
            elif choice == '5':
                spinner.stop()
                user_location = input(
                    'Enter your postcode or specific location: ')
                spinner = Spinner("Sorting events...")
                spinner.start()
                api_key = os.getenv('GOOGLE_MAPS_API_KEY')
                closest_events = find_closest_events(user_location,
                                                     events, api_key)
                # This is calculated using the geopy library's geodesic
                # function, which calculates the distance between two points
                # on the Earth's surface using the geodesic distance,
                # which is more accurate than the haversine formula
                if closest_events:
                    display_events(closest_events, 0, len(closest_events),
                                   'data-manipulation', 'None')
                else:
                    print('\n-------------------------------------'
                          '\nNo events found or location cannot be geocoded.'
                          '\n-------------------------------------')
                    continue
                del user_location  # Ensure the user_location is deleted
            elif choice == '6':
                spinner.stop()
                print('\n-------------------------------------'
                      '\nReturning to the main menu.'
                      '\n-------------------------------------')
                main()
                return
            else:
                print('\n-------------------------------------'
                      '\nInvalid choice. Please try again.'
                      '\n-------------------------------------')
                spinner.stop()
                continue
        finally:
            spinner.stop()


def event_manipulation_menu(events):
    """
    A menu to manipulate the collected events data the user wants to view.

    Args:
        events (list of dict): The user's collected events.
    """
    while True:
        print('\nChoose an option to manipulate the events data:')
        print('1. Sort events')
        print('2. Compare events')
        print('3. Main Menu')
        choice = input('Enter your choice: ').strip()

        if choice == '1':
            sort_events(events)
        elif choice == '2':
            compare_events(events)
        elif choice == '3':
            print('Returning to the main menu.')
            main()
            return
        else:
            print('\n-------------------------------------'
                  '\nInvalid choice. Please try again.'
                  '\n-------------------------------------')


def get_unique_search_keys():
    """
    This function provides a list of all the users searches,
    when the user searches for events, the search key is stored
    along with each event in the collection. This works like a
    class to identify the events under groups of search keys.

    Returns:
        List of strings: A list of all the unique search keys
                        in the collection.
    """
    pipeline = [
        {'$group': {'_id': '$search_key'}},
    ]
    # A pipeline is a list of operations that MongoDB
    # applies to the documents in a collection,
    # the $group operator groups documents by a specified expression,
    # _id is a system variable that holds the value of the field
    # that the $group operator groups by, in this case, the search_key field,
    # without the $ mongodb would treat it as a litteral string
    # rather than a reference to a field
    unique_search_keys = list(collection.aggregate(pipeline))
    # Make a list of all the unique search keys in the collection (mongo db)
    # and aggregate it against the pipeline
    # example of the pipeline output: [{'_id': 'data-manipulation_london'},
    # {'_id': 'data-manipulation_birmingham'}]
    return [key['_id'] for key in unique_search_keys]
    # For every key in the unique_search_keys list, return the value
    # of the item with its '_id' field


def search_events_in_collection():
    """
    Uses the return from get_unique_search_keys function to
    display the users searches and allow the user to select
    a search key to view the events for that search key.

    Returns:
        List of dict: A list of all the events under the selected search key.
    """
    unique_search_keys = get_unique_search_keys()
    user_selection = 'data-manipulation'

    if len(unique_search_keys) == 0:
        print("No events found in the collection.")
        return

    print("-------------------------------------\nChoose a search key:")
    for i, key in enumerate(unique_search_keys, 1):
        # Enumerate will number each unique search key starting from 1,
        # 'i' will be the number and 'key' will be the search key
        print(f"{i}. {key}")

    while True:
        try:
            choice = input("\nEnter the number of your choice: ").strip()
            choice_index = int(choice) - 1
            # Ensure choice is an interger and zero-indexed
            if choice_index < 0 or choice_index >= len(unique_search_keys):
                # Check if the choice is within the range
                # of the unique search keys
                raise IndexError
                # Raise an IndexError if the choice is out of range
            all_events = list(collection.find(
                {'search_key': unique_search_keys[choice_index]}))
            # Make a list of all mongodb data with the search_key that
            # matches the user choice
            break
        except ValueError:
            print('\n-------------------------------------'
                  '\nInvalid input. Please enter a valid number.'
                  '\n-------------------------------------')
        except IndexError:
            print(f'\n-------------------------------------'
                  f'\nInvalid choice. Please enter a number'
                  f' between 1 and {len(unique_search_keys)}.'
                  f'\n-------------------------------------')

    if not all_events:
        print("No events found for the selected search key.")
        return

    display_events(all_events, 0, len(all_events), user_selection,
                   search_key=unique_search_keys[int(choice) - 1])
    while True:
        save_choice = input('-------------------------------------'
                            '\nWould you like to save the events to a CSV or'
                            ' Excel file? (C/E)\nOr perform tasks'
                            ' on the data? (T): ').strip().lower()
        if save_choice == 'c':
            try:
                save_to_csv(all_events)
                return
            except ValueError as e:
                print(f'Error saving events to CSV: {e}')
        elif save_choice == 'e':
            try:
                save_to_excel(all_events)
                return
            except ValueError as e:
                print(f'Error saving events to Excel: {e}')
        elif save_choice == 't':
            event_manipulation_menu(all_events)
        else:
            print('\n-------------------------------------'
                  '\nInvalid choice. Please C to print to CSV, E to print to'
                  ' Excel\nOr T to perform tasks with the event data.'
                  '\n-------------------------------------')
            continue


def view_all_events():
    """
    This function displays all the events in the collection
    with the option to save the events to a CSV, Excel file or
    perform tasks on the data.
    """
    all_events = list(collection.find({}))
    user_selection = 'data-manipulation'
    if len(all_events) == 0:
        print('No events found')
        return

    display_events(all_events, 0, len(all_events),
                   user_selection, search_key='None')
    while True:
        save_choice = input('-------------------------------------'
                            '\nWould you like to save the events to a CSV or'
                            ' Excel file? (C/E)\nOr perform tasks on'
                            ' the data? (T): ').strip().lower()
        if save_choice == 'c':
            try:
                save_to_csv(all_events)
                return
            except ValueError as e:
                print(f"Error saving events to CSV: {e}")
        elif save_choice == 'e':
            try:
                save_to_excel(all_events)
                return
            except ValueError as e:
                print(f"Error saving events to Excel: {e}")
        elif save_choice == 't':
            event_manipulation_menu(all_events)
        else:
            print('\n-------------------------------------'
                  '\nInvalid choice. Please C to print to CSV, E to print to'
                  ' Excel\nOr T to perform tasks with the event data.'
                  '\n-------------------------------------')
            continue


def search_events():
    """
    Option 1 on the main menu, allows the user to search for events
    with optional event type, location parameters and date range.

    Returns:
        List of dict: A list of scraped events from the selected search,
                    and optional location and date range.
    """
    product = input('Enter event type or name: ').replace(' ', '%20')
    location = input('Enter location: ').replace(' ', '%20')
    print('Would you like to enter a date? (Y/N)')
    date_choice = input('Enter your choice: ').strip().lower()
    start_date = ''
    end_date = ''
    day = ''
    tags_counter = Counter()

    if date_choice == 'y':
        print('Please enter an option: ')
        print('1. Today')
        print('2. Tomorrow')
        print('3. This weekend')
        print('4. Pick a date')
        day = input('Enter the number of choice: ')
        if day == '1':
            day = 'today'
        elif day == '2':
            day = 'tomorrow'
        elif day == '3':
            day = 'this-weekend'
        else:
            start_date = input('Enter the start date (YYYY-MM-DD): ')
            end_date = input('Enter the end date (YYYY-MM-DD): ')

    search_key = f'{product}_{location}'
    user_selection = 'eventbrite'
    spinner = Spinner("Fetching events...")
    spinner.start()
    unique_events = []
    page_number = 1

    try:
        # Check if the search term is in the cache
        if search_key in cache:
            spinner.stop()
            print("Using cached events from hashtable.")
            unique_events = cache[search_key]
        else:
            spinner.stop()
            print('Loading times may vary depending on'
                  ' the scope of the search.'
                  '\nBroader searches may take longer to load.'
                  ' Avrg time: 20sec-3min')
            spinner = Spinner("Scraping new events...")
            spinner.start()
            events_data, tags_counter, _ = scrape_eventbrite_events(
                location, day, product, page_number, start_date, end_date)
            unique_events.extend(events_data)
            cache[search_key] = unique_events
    finally:
        spinner.stop()

    result = display_paginated_events(
        unique_events, search_key, user_selection, location, day,
        start_date, end_date, tags_counter, product, page_number)

    if result == 'new_search':
        main()
        return


def search_top_categories():
    """
    Option 3 on the main menu allows the user to search for top set
    categories of events with a optional location and date range parameters.

    Returns:
        List of dict: A list of scraped events from the selected category,
                    and optional location and date range.
    """
    categories = [
        'Home & Lifestyle', 'Business', 'Health', 'Performing & Visual Arts',
        'Family & Education', 'Holidays', 'Music', 'Community',
        'Hobbies', 'Charity & Causes', 'Food & Drink', 'Science & Tech',
        'Sports & Fitness', 'Travel & Outdoor', 'Spirituality', 'Nightlife',
        'Dating', 'Film & Media', 'Fashion', 'Government', 'Auto, Boat & Air',
        'School Activities'
    ]

    user_selection = 'eventbrite_top'
    product = None
    tags_counter = Counter()

    def display_categories():
        print('\nPlease choose a category:')
        for i, category in enumerate(categories, 1):
            print(f'{i}. {category}')

    def get_user_choice():
        while True:
            user_input = input(f'Enter the number of your'
                               f' choice to find events in {location}: ')
            try:
                choice = int(user_input)
                if 1 <= choice <= len(categories):
                    return categories[choice - 1]
                # 1 is subtracted from the choice to get the correct index
                else:
                    print(f'Please enter a number between 1 and {len(
                        categories)}.')
            except ValueError:
                print('-------------------------------------'
                      '\nInvalid input. Please enter a number.'
                      '\n-------------------------------------')

    def generate_slug(category):
        category = category.replace('&', 'and')
        return re.sub(r'\s+', '-', category.strip().lower())

    country = 'united-kingdom'
    location = input('Enter location: ').replace(' ', '')
    print('Would you like to enter a date? (Y/N)')
    date_choice = input('Enter your choice: ').strip().lower()
    start_date = ''
    end_date = ''
    day = ''

    if date_choice == 'y':
        print('Please enter an option: ')
        print('1. Today')
        print('2. Tomorrow')
        print('3. This weekend')
        print('4. Pick a date')
        day = input('Enter the number of choice: ')
        if day == '1':
            day = 'today'
        elif day == '2':
            day = 'tomorrow'
        elif day == '3':
            day = 'this-weekend'
        else:
            start_date = input('Enter the start date (YYYY-MM-DD): ')
            end_date = input('Enter the end date (YYYY-MM-DD): ')

    display_categories()
    category = get_user_choice()
    search_key = f'{generate_slug(category)}_{location}_{country}'
    spinner = Spinner("Fetching events...")
    spinner.start()
    category_slug = generate_slug(category)
    unique_events = []
    page_number = 1

    try:
        # Check if the search term is in the cache
        if search_key in cache:
            spinner.stop()
            print("Using cached events from hashtable.")
            unique_events = cache[search_key]
        else:
            spinner.stop()
            print('Loading times may vary depending on'
                  ' the scope of the search.'
                  '\nBroader searches may take longer to load.'
                  ' Avrg time: 20sec-3min')
            spinner = Spinner("Scraping new events...")
            spinner.start()
            events_data, tags_counter, _ = scrape_eventbrite_categories(
                location, category_slug, day, page_number,
                start_date, end_date)
            unique_events.extend(events_data)
            cache[search_key] = unique_events
    finally:
        spinner.stop()

    result = display_paginated_events(
        unique_events, search_key, user_selection, location, day,
        start_date, end_date, tags_counter, category_slug, product,
        page_number)

    if result == 'new_search':
        main()
        return


def search_top_events():
    """
    Option 2 on the main menu, allows the user to search for top events
    with an optional location parameter.

    Returns:
        List of dict: A list of scraped top events from the optional location,
                    or a list of scraped top events from the UK.
    """
    location = input('Enter location: ').replace(' ', '')
    search_key = f'all_top_events_{location}'
    spinner = Spinner("Fetching events...")
    spinner.start()
    unique_events = []
    user_selection = 'eventbrite_top_no_category'
    day = ''
    start_date = ''
    end_date = ''
    tags_counter = Counter()

    try:
        if search_key in cache:
            spinner.stop()
            print("Using cached events from hashtable.")
            unique_events = cache[search_key]
        else:
            spinner.stop()
            print('Loading times may vary depending on'
                  ' the scope of the search.'
                  '\nBroader searches may take longer to load.'
                  ' Avrg time: 20sec-3min')
            spinner = Spinner("Scraping new events...")
            spinner.start()
            events_data, tags_counter, _ = scrape_eventbrite_top_events(
                location)
            unique_events.extend(events_data)
            cache[search_key] = unique_events
    finally:
        spinner.stop()

    result = display_paginated_events(
        unique_events, search_key, user_selection, location,
        day, start_date, end_date, tags_counter)
    if result == 'new_search':
        main()
        return


def display_common_tags(tags_counter):
    """
    Display the most common tags from the collected events.

    Args:
        tags_counter (Counter): A Counter object containing
                            a count of unique tags

    Returns:
        List: A list of tuples containing the most common tags
        and their counts.
    """
    if tags_counter:
        spinner = Spinner("Loading most common tags...")
        spinner.start()
        most_common_tags = tags_counter.most_common(6)
        spinner.stop()
        print('\nThe most common tags are:')
        for tag, count in most_common_tags:
            print(f'{tag}: {count}')


def main():
    """
    The first function to run when the program starts.
    Functions as the main menu for the program.
    """
    while True:
        print('-------------------------------------'
              '\nWelcome to Event Hoarder!\nSearch for events and they will be'
              ' automatically be saved to a database so you '
              '\ncan perform sorting, comparing or filtering tasks,'
              ' also print to Excel or CSV'
              '\nSearch for events with either of the first 3 options,'
              '\nAfter viewing your searched events, you can view them'
              ' in the database\nwith option 4 or quickly access within'
              ' searches with "T" \nOption 5 is for viewing links to saved Excel,'
              ' CSV or data visuals that\nyou may make.'
              '\n-------------------------------------')
        print('\nChoose an option:')
        print('1. Quick Search & Collect')
        print('2. Search & Collect Top Events')
        print('3. Search & Collect Top Categories')
        print('4. View Collected Events')
        print('5. View links for saved Excel , CSV or data visuals')
        print('6. Exit')
        print('#. Clear Database')
        choice = input("Enter your choice: ").strip()

        if choice == '1':
            search_events()
        elif choice == '2':
            search_top_events()
        elif choice == '3':
            search_top_categories()
        elif choice == '4':
            collection_menu()
        elif choice == '5':
            view_data_files()
        elif choice == '6':
            leave = input('Are you sure you want to exit? Exiting will delete'
                          ' any files you may have made.\nBe sure to'
                          ' view/download them first! (Y/N): ').strip().lower()
            if leave == 'y':
                delete_all_files_in_gcs('data-visuals-serving')
                time.sleep(3)  # Ensure deletion process has time to complete
                print('-------------------------------------'
                      '\nExiting the program'
                      '\n-------------------------------------.')
                sys.exit()
            else:
                continue
        elif choice == '#':
            collection.delete_many({})
            print('-------------------------------------'
                  '\nDatabase cleared\n-------------------------------------.')
            main()
            return
        else:
            print('\n-------------------------------------'
                  '\nInvalid choice. Please try again.'
                  '\n-------------------------------------')


def display_paginated_events(unique_events, search_key, user_selection,
                             location, day, start_date, end_date, tags_counter,
                             category_slug=None, product=None, page_number=1):
    """
    Display the scraped events in a paginated format with 5 per
    console page and fetch more events if requested and available.

    Args:
       - unique_events (list of dict): A list of unique scraped events.
       - search_key (str): The user given search term.
       - user_selection (str): A string used in backend functions
                                to identify the user's choice.
       - location (str, optional): The user given location.
       - day (str, optional): The user given day.
       - start_date (str, optional): The user given event start date.
       - end_date (str, optional): The user given event end date.
       - tags_counter (Counter): A Counter object containing a count
                                of unique tags.
       - category_slug (str, optional): The category slug for the top
                                        categories search. Defaults to None.
       - product (str, optional): The users given event type.
                                    Defaults to None.
       - page_number (int, optional): Page number used for URL pagination.
                                    Defaults to 1.

    Returns:
        str: 'done' if the user is done viewing events or no more events,
        'new_search' if the user wants to start a new search.
    """
    page_size = 5
    total_events = len(unique_events)
    current_page = 0
    more_events_check = True
    should_break = False

    while more_events_check:
        start_index = current_page * page_size
        # The start index is the current page number multiplied by the page
        # size (number of events per console page)
        end_index = min(start_index + page_size, total_events)
        # The end index is the minimum of the start index plus the page size
        # and the total number of events
        display_events(unique_events, start_index,
                       end_index, user_selection, search_key)

        if end_index >= total_events:
            # Fetch more events if available
            page_number += 1
            print('Loading times may vary depending on'
                  ' the scope of the search.'
                  '\nBroader searches may take longer to load.'
                  ' Avrg time: 20sec-3min')
            spinner = Spinner("Fetching more events...")
            spinner.start()
            try:
                if user_selection == 'eventbrite':
                    events_data, new_tags_counter, more_events_check = \
                        scrape_eventbrite_events(
                            location, day, product,
                            page_number, start_date, end_date)
                elif user_selection == 'eventbrite_top':
                    events_data, new_tags_counter, more_events_check = \
                        scrape_eventbrite_categories(
                            location, category_slug, day,
                            page_number, start_date, end_date)
                elif user_selection == 'eventbrite_top_no_category':
                    events_data, new_tags_counter, more_events_check = \
                        scrape_eventbrite_top_events(location)
                else:
                    break

                print(f'Fetched {len(events_data)}'
                      f' events for page number: {page_number} ')
                unique_events.extend(events_data)
                # Add the new events to the unique events list
                tags_counter.update(new_tags_counter)
                # Update the tags counter with the new tags
                total_events = len(unique_events)
                # Update the total number of events
                print(f'Total events after fetching: {total_events}'
                      '\nNext page of events has been fetched.'
                      '\nYou may press "Y" and view more events.')
                # Display the most common tags
                display_common_tags(tags_counter)
            finally:
                spinner.stop()
                if not more_events_check:
                    print("No more events to fetch.")
                    should_break = True

            if should_break:
                break

        user_input = input('-------------------------------------'
                           '\nPress "Y" to see more events, "T" to quickly go'
                           ' and view/manipulate this data'
                           ' "S" go back to main menu, or'
                           ' any other key to exit: ').strip().lower()
        if user_input == 's':
            return 'new_search'
        elif user_input == 't':
            collection_menu()
        elif user_input == 'y':
            current_page += 1
        else:
            print("Exiting the program.")
            sys.exit()

    return 'done'


if __name__ == "__main__":
    main()
